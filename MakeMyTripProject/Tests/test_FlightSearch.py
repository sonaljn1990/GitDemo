import unittest
import time
import openpyxl
workbook = openpyxl.load_workbook("D:/Python practice files/MMT.xlsx")
sheet = workbook.active
from selenium import webdriver
from MakeMyTripProject.Pages.FlightBookingPage import FlightBooking

class test_flightSearch(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(executable_path="C:\Drivers\chromedriver_win32\chromedriver.exe")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_03_flight_page_title(self):
        driver = self.driver
        driver.get("https://www.makemytrip.com/")
        time.sleep(3)
        print("GitHub command to check")
        titleOfTheWebPage = driver.title
        if titleOfTheWebPage == "MakeMyTrip - #1 Travel Website 50% OFF on Hotels, Flights & Holiday":
            print("test 03 is pass")
            sheet.cell(row=4, column=4).value = 'Pass'
            workbook.save("D:/Python practice files/MMT.xlsx")
        else:
            print("test 03 is fail")
            assert titleOfTheWebPage == "MakeMyTrip - #1 Travel Website 50% OFF on Hotels, Flights & Holiday", "You are on wrong page"
            sheet.cell(row=4, column=4).value = 'Fail'
            workbook.save("D:/Python practice files/MMT.xlsx")


    def test_04_flight_search(self):
        driver = self.driver
        driver.get("https://www.makemytrip.com/")
        flightSearch = FlightBooking(driver)
        driver.implicitly_wait(5)
        time.sleep(3)
        flightSearch.select_round_trip_option()
        time.sleep(3)
        flightSearch.from_city = sheet.cell(row=5,column=2).value
        time.sleep(2)
        flightSearch.select_from_city(flightSearch.from_city)
        time.sleep(2)
        flightSearch.to_city = sheet.cell(row=5, column=3).value
        time.sleep(2)
        flightSearch.select_to_city(flightSearch.to_city)
        time.sleep(2)
        flightSearch.select_departure_date()
        time.sleep(2)
        flightSearch.select_return_date()
        time.sleep(2)
        flightSearch.select_travellers_information()
        time.sleep(2)
        flightSearch.perform_search_for_flights()
        print("GitHub command to check")
        print("GitHub command to check")
        print("GitHub command to check")


    @classmethod
    def tearDownClass(cls):
        cls.driver.close()


workbook.save("D:/Python practice files/MMT.xlsx")

if __name__ == "__main__":
    unittest.main()
