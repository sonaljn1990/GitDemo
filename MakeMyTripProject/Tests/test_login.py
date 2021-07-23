from selenium import webdriver
import MakeMyTripProject.Pages.LoginPage
import unittest
import time
import openpyxl

workbook = openpyxl.load_workbook("D:/Python practice files/MMT.xlsx")
sheet = workbook.active

class test_login(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(executable_path="C:\Drivers\chromedriver_win32\chromedriver.exe")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_01_valid_login(self):
        driver = self.driver
        driver.get("https://www.makemytrip.com/")
        login = MakeMyTripProject.Pages.LoginPage.LoginPage(driver)
        login.click_to_login_via_email()
        time.sleep(2)
        login.email = sheet.cell(row=2, column=2).value
        login.password = sheet.cell(row=2, column=3).value
        login.enter_username(login.email)
        login.click_on_continue_button()
        login.enter_password(login.password)
        login.click_login_button()
        time.sleep(5)
        if login.check_user_loggedIn():
            print("test 01 is pass")
            sheet.cell(row=2, column=4).value = 'Pass'
            workbook.save("D:/Python practice files/MMT.xlsx")
            login.click_logged_in_user()
            login.click_on_my_profile()
            login.click_on_logout()
        else:
            print("test 01 is fail")
            sheet.cell(row=2, column=4).value = 'Fail'
            workbook.save("D:/Python practice files/MMT.xlsx")

    def test_02_invalid_login(self):
        driver = self.driver
        driver.get("https://www.makemytrip.com/")
        login = MakeMyTripProject.Pages.LoginPage.LoginPage(driver)
        login.click_on_Login_or_create_account()
        time.sleep(2)
        login.email = sheet.cell(row=3, column=2).value
        login.password = sheet.cell(row=3, column=3).value
        login.enter_username(login.email)
        login.click_on_continue_button()
        login.enter_password(login.password)
        login.click_login_button()
        time.sleep(3)
        if login.check_invalid_credential_login_msg():
            print("test 02 is pass, user is not able to login with invalid credentials")
            sheet.cell(row=3, column=4).value = 'Pass'
            workbook.save("D:/Python practice files/MMT.xlsx")
            login.driver.refresh()
        else:
            print("something went wrong")
            sheet.cell(row=3, column=4).value = 'Fail'
            workbook.save("D:/Python practice files/MMT.xlsx")

    #@classmethod
    #def tearDownClass(cls):
        #cls.driver.close()


workbook.save("D:/Python practice files/MMT.xlsx")

if __name__ == "__main__":
    unittest.main()
